"""
run_roi_partial_rsa_exploratory.py
==================================
ROI-based partial RSA — EXPLORATORY ANALYSIS

All time windows from 100-200 ms onward (excluding the hypothesis-driven 50-100 ms).
Extended set of ROIs based on Ntoumanis et al. (2026) contrasts.
FDR correction across ROIs × time windows.

Author: Ioannis Ntoumanis
Created: 2026
"""

import numpy as np
import nibabel as nib
from scipy.spatial.distance import squareform
from scipy.stats import rankdata
import matplotlib.pyplot as plt
import h5py
import os
import pickle
import time

# ===========================================================================
# CONFIGURATION
# ===========================================================================

SOURCE_AVG_FILE = r"Source data/source_avg.npy"
VALID_MASK_FILE = r"Source data/valid_mask.npy"
GRID_LOC_FILE   = r"Source data/grid_loc.npy"
MNI_LOC_FILE    = r"Source data/grid_loc_mni.mat"
TIME_WIN_FILE   = r"Source data/time_windows.npy"
MODEL_RDM_DIR   = r"Results/Model RDMs"
AAL_ATLAS_FILE  = r"Code/AAL3v1.nii"
OUTPUT_DIR      = r"Results/Late effects"

MODEL_NAMES  = ["rdm_valence", "rdm_arousal"]
MODEL_LABELS = ["Valence", "Arousal"]

N_PERM       = 10000
RANDOM_SEED  = 25
FDR_Q        = 0.05
MIN_SUBJECTS = 15
ASSIGNMENT_RADIUS = 10  # mm

# Skip the hypothesis-driven window (50-100 ms = index 0)
SKIP_TIME_IDX = 0

ROIS = {
    # Temporal
    'Angular':      ['Angular_L', 'Angular_R'],
    'MTG':          ['Temporal_Mid_L',     'Temporal_Mid_R'],
    
    # Occipital / Ventral visual
    'IOG':          ['Occipital_Inf_L', 'Occipital_Inf_R'],
    'Lingual':      ['Lingual_L', 'Lingual_R'],
    'Fusiform':     ['Fusiform_L', 'Fusiform_R'],

    # Medial / Cingulate
    'PCC':          ['Cingulate_Post_L', 'Cingulate_Post_R'],
    'Precuneus':    ['Precuneus_L', 'Precuneus_R'],
    'MCC':          ['Cingulate_Mid_L', 'Cingulate_Mid_R'],
    'ACC':          ['ACC_sup_L', 'ACC_sup_R'],

    # Subcortical
    'Amygdala':     ['Amygdala_L', 'Amygdala_R'],
    'Hippocampus':  ['Hippocampus_L', 'Hippocampus_R'],

    # Frontal
    'dlPFC':        ['Frontal_Sup_2_L',    'Frontal_Sup_2_R'],
    'MFG':          ['Frontal_Mid_2_L', 'Frontal_Mid_2_R'],
    'IFG':          ['Frontal_Inf_Oper_L', 'Frontal_Inf_Oper_R'],
    'OFC':          ['Frontal_Med_Orb_L', 'Frontal_Med_Orb_R']
}

# ===========================================================================
# HELPER FUNCTIONS
# ===========================================================================

def rank_vector(v):
    return rankdata(v).astype(np.float64)

def fdr_bh(p_values, q=0.05):
    """Benjamini-Hochberg FDR on a flat array of p-values."""
    p_flat = p_values.flatten()
    valid = ~np.isnan(p_flat)
    p_valid = p_flat[valid]
    n = len(p_valid)
    if n == 0:
        return np.zeros_like(p_values, dtype=bool).reshape(p_values.shape), 0.0

    sorted_idx = np.argsort(p_valid)
    sorted_p = p_valid[sorted_idx]
    bh_critical = np.arange(1, n + 1) / n * q
    below = sorted_p <= bh_critical

    if not np.any(below):
        return np.zeros_like(p_values, dtype=bool).reshape(p_values.shape), 0.0

    threshold = sorted_p[np.max(np.where(below)[0])]
    significant = np.zeros_like(p_flat, dtype=bool)
    significant[valid] = p_flat[valid] <= threshold
    return significant.reshape(p_values.shape), threshold

def assign_grid_to_aal(grid_mm, aal_img, aal_labels, aal_indices, roi_dict,
                        radius_mm=7.5):
    """Assign each grid point to an AAL region based on nearest labeled voxel."""
    aal_data = aal_img.get_fdata()
    aal_affine = aal_img.affine
    inv_affine = np.linalg.inv(aal_affine)

    index_to_roi = {}
    for roi_name, label_substrings in roi_dict.items():
        for substr in label_substrings:
            for i, label in enumerate(aal_labels):
                if label == substr or substr in label:
                    idx = int(aal_indices[i])
                    index_to_roi[idx] = roi_name

    print(f"  AAL indices mapped to ROIs: {len(index_to_roi)}")

    roi_assignments = {name: [] for name in roi_dict.keys()}

    for g in range(len(grid_mm)):
        mni_coord = np.array([grid_mm[g, 0], grid_mm[g, 1], grid_mm[g, 2], 1.0])
        vox = inv_affine @ mni_coord
        vi, vj, vk = int(round(vox[0])), int(round(vox[1])), int(round(vox[2]))

        if (0 <= vi < aal_data.shape[0] and
            0 <= vj < aal_data.shape[1] and
            0 <= vk < aal_data.shape[2]):
            aal_idx = int(aal_data[vi, vj, vk])
            if aal_idx in index_to_roi:
                roi_assignments[index_to_roi[aal_idx]].append(g)

        if not assigned:
            search_range = int(np.ceil(radius_mm / np.abs(aal_affine[0, 0])))
            found = False
            for di in range(-search_range, search_range + 1):
                if found: break
                for dj in range(-search_range, search_range + 1):
                    if found: break
                    for dk in range(-search_range, search_range + 1):
                        ni, nj, nk = vi + di, vj + dj, vk + dk
                        if (0 <= ni < aal_data.shape[0] and
                            0 <= nj < aal_data.shape[1] and
                            0 <= nk < aal_data.shape[2]):
                            aal_idx = int(aal_data[ni, nj, nk])
                            if aal_idx in index_to_roi:
                                vox_mm = aal_affine @ np.array([ni, nj, nk, 1.0])
                                dist   = np.sqrt(np.sum((grid_mm[g] - vox_mm[:3])**2))
                                if dist <= radius_mm:
                                    roi_assignments[index_to_roi[aal_idx]].append(g)
                                    found = True
                                    break

    for name in roi_assignments:
        roi_assignments[name] = np.array(roi_assignments[name], dtype=int)
        print(f"    {name}: {len(roi_assignments[name])} grid points")

    return roi_assignments

def make_nifti(grid_mm, values, resolution=5):
    """Create a NIfTI volume from scattered MNI points."""
    x_range = np.arange(-80, 85, resolution)
    y_range = np.arange(-115, 80, resolution)
    z_range = np.arange(-65, 90, resolution)

    vol = np.zeros((len(x_range), len(y_range), len(z_range)))

    for i in range(len(grid_mm)):
        if values[i] == 0: continue
        xi = np.argmin(np.abs(x_range - grid_mm[i, 0]))
        yi = np.argmin(np.abs(y_range - grid_mm[i, 1]))
        zi = np.argmin(np.abs(z_range - grid_mm[i, 2]))
        vol[xi, yi, zi] = values[i]

    affine = np.eye(4)
    affine[0, 3] = x_range[0]
    affine[1, 3] = y_range[0]
    affine[2, 3] = z_range[0]
    affine[0, 0] = resolution
    affine[1, 1] = resolution
    affine[2, 2] = resolution

    return nib.Nifti1Image(vol, affine)

# ===========================================================================
# MAIN
# ===========================================================================
t_start = time.time()
os.makedirs(OUTPUT_DIR, exist_ok=True)

print("=" * 60)
print("ROI PARTIAL RSA — EXPLORATORY (100-1000 ms)")
print(f"  Permutations: {N_PERM}")
print("=" * 60)

# ----- 1. Load data -----
print("\nStep 1: Loading data")
source_avg   = np.load(SOURCE_AVG_FILE)
valid_mask   = np.load(VALID_MASK_FILE)
grid_coords  = np.load(GRID_LOC_FILE)
time_windows_all = np.load(TIME_WIN_FILE)

n_stimuli, n_grid, n_time_all = source_avg.shape
valid_idx = np.where(valid_mask)[0]
n_valid = len(valid_idx)

# Drop the early window
time_mask = np.arange(n_time_all) != SKIP_TIME_IDX
time_indices = np.where(time_mask)[0]
time_windows = time_windows_all[time_indices]
n_time = len(time_indices)

print(f"  Source: {source_avg.shape}")
print(f"  Valid stimuli: {n_valid}")
print(f"  Time windows (exploratory): {n_time} "
      f"({time_windows[0,0]*1000:.0f}–{time_windows[-1,1]*1000:.0f} ms)")
print(f"  Skipped: window {SKIP_TIME_IDX} "
      f"({time_windows_all[SKIP_TIME_IDX,0]*1000:.0f}–"
      f"{time_windows_all[SKIP_TIME_IDX,1]*1000:.0f} ms, hypothesis-driven)")

with h5py.File(MNI_LOC_FILE, 'r') as f:
    grid_mni = np.array(f['grid_mni']).T
grid_mm = grid_mni * 1000

n_models = len(MODEL_NAMES)
model_rdms_trimmed = []
for name in MODEL_NAMES:
    rdm = np.load(os.path.join(MODEL_RDM_DIR, f"{name}.npy"))
    rdm_trim = rdm[np.ix_(valid_idx, valid_idx)]
    model_rdms_trimmed.append(rdm_trim)

n_pairs = n_valid * (n_valid - 1) // 2

# ----- 2. Build design matrices -----
print("\nStep 2: Building design matrices")
obs_model_vecs = []
for m_idx in range(n_models):
    vec = squareform(model_rdms_trimmed[m_idx], checks=False)
    rv = rank_vector(vec)
    rv -= rv.mean()
    obs_model_vecs.append(rv)

X_obs = np.column_stack(obs_model_vecs + [np.ones(n_pairs)])
Xt_obs = X_obs.T
XtX_inv_obs = np.linalg.inv(Xt_obs @ X_obs)

print(f"  Design matrix: {X_obs.shape}")

print(f"  Building {N_PERM} permuted design matrices...", end=" ", flush=True)
rng = np.random.default_rng(RANDOM_SEED)
perm_orders = np.zeros((N_PERM, n_valid), dtype=np.int64)
for p in range(N_PERM):
    perm_orders[p] = rng.permutation(n_valid)

Xt_stack = np.zeros((N_PERM, n_models + 1, n_pairs), dtype=np.float64)
XtX_inv_stack = np.zeros((N_PERM, n_models + 1, n_models + 1), dtype=np.float64)

for p in range(N_PERM):
    cols = []
    for m_idx in range(n_models):
        rdm_perm = model_rdms_trimmed[m_idx][np.ix_(perm_orders[p], perm_orders[p])]
        vec = rank_vector(squareform(rdm_perm, checks=False))
        vec -= vec.mean()
        cols.append(vec)
    cols.append(np.ones(n_pairs))
    X_perm = np.column_stack(cols)
    Xt_stack[p] = X_perm.T
    XtX_inv_stack[p] = np.linalg.inv(X_perm.T @ X_perm)
print("done")

# ----- 3. Assign grid points to ROIs -----
print("\nStep 3: Assigning grid points to AAL3 ROIs")
aal_img = nib.load(AAL_ATLAS_FILE)

aal_dir = os.path.dirname(AAL_ATLAS_FILE)
labels = []; indices = []
for label_file in ['AAL3v1.nii.txt', 'AAL3v1_1mm.nii.txt', 'ROI_MNI_V7.txt',
                   'AAL3v1.xml', 'AAL3_labels.txt']:
    lf = os.path.join(aal_dir, label_file)
    if os.path.exists(lf):
        print(f"  Found label file: {lf}")
        with open(lf, 'r') as f:
            for line in f:
                parts = line.strip().split()
                if len(parts) >= 2:
                    try:
                        idx = int(parts[0])
                        name = parts[1]
                        indices.append(idx)
                        labels.append(name)
                    except ValueError:
                        continue
        break

roi_assignments = assign_grid_to_aal(grid_mm, aal_img, labels, indices, ROIS,
                                      radius_mm=ASSIGNMENT_RADIUS)

# Minimum 5 grid points per ROI
active_rois = {name: idx for name, idx in roi_assignments.items() if len(idx) >= 5}
excluded_rois = {name: len(idx) for name, idx in roi_assignments.items() if len(idx) < 5}
roi_names = sorted(active_rois.keys())
n_rois = len(roi_names)

print(f"\n  Active ROIs (>= 5 grid points): {n_rois}")
if excluded_rois:
    print(f"  Excluded:")
    for name, n_pts in excluded_rois.items():
        print(f"    {name}: {n_pts} grid points")

# ----- 4. ROI partial RSA + permutations -----
print(f"\nStep 4: Exploratory ROI partial RSA "
      f"({n_rois} ROIs × {n_time} time windows × {N_PERM} permutations)")
print("=" * 60)

source_valid = source_avg[valid_idx]

roi_observed   = np.zeros((n_rois, n_time, n_models), dtype=np.float64)
roi_pvalues    = np.ones((n_rois, n_time, n_models), dtype=np.float64)
roi_perm_betas = np.zeros((n_rois, n_time, n_models, N_PERM), dtype=np.float32)

for r_idx, roi_name in enumerate(roi_names):
    grid_indices = active_rois[roi_name]
    n_roi_points = len(grid_indices)
    print(f"\n  {roi_name} ({n_roi_points} grid points)")

    for t_local, t_global in enumerate(time_indices):
        patterns = source_valid[:, grid_indices, t_global]

        patterns_zm = patterns - patterns.mean(axis=1, keepdims=True)
        norms = np.sqrt(np.sum(patterns_zm ** 2, axis=1))

        if np.any(norms < 1e-12):
            roi_observed[r_idx, t_local] = np.nan
            roi_pvalues[r_idx, t_local] = np.nan
            continue

        patterns_normed = patterns_zm / norms[:, np.newaxis]
        corr_matrix = patterns_normed @ patterns_normed.T
        np.clip(corr_matrix, -1.0, 1.0, out=corr_matrix)
        neural_rdm = 1.0 - corr_matrix
        neural_vec = squareform(neural_rdm, checks=False)
        neural_ranked = rank_vector(neural_vec)
        neural_ranked -= neural_ranked.mean()

        obs_betas = XtX_inv_obs @ (Xt_obs @ neural_ranked)
        for m_idx in range(n_models):
            roi_observed[r_idx, t_local, m_idx] = obs_betas[m_idx]

        XtY_all = np.einsum('ijk,k->ij', Xt_stack, neural_ranked)
        perm_betas_all = np.einsum('ijk,ik->ij', XtX_inv_stack, XtY_all)
        
        roi_perm_betas[r_idx, t_local, :, :] = perm_betas_all[:, :n_models].T

        for m_idx in range(n_models):
            roi_pvalues[r_idx, t_local, m_idx] = np.mean(
                perm_betas_all[:, m_idx] >= obs_betas[m_idx])

    for m_idx in range(n_models):
        valid_p = roi_pvalues[r_idx, :, m_idx]
        valid_p = valid_p[~np.isnan(valid_p)]
        if len(valid_p) > 0:
            min_p = valid_p.min()
            best_t = np.nanargmin(roi_pvalues[r_idx, :, m_idx])
            best_beta = roi_observed[r_idx, best_t, m_idx]
            tw_s = time_windows[best_t, 0] * 1000
            tw_e = time_windows[best_t, 1] * 1000
            sig_marker = " *" if min_p < 0.05 else ""
            print(f"    {MODEL_LABELS[m_idx]}: best p={min_p:.4f} "
                  f"(beta={best_beta:.4f}, {tw_s:.0f}–{tw_e:.0f} ms){sig_marker}")

# ----- 5. FDR correction across ROIs × time windows -----
print("\n" + "=" * 60)
print("Step 5: FDR correction across ROIs × time windows")
total_tests = n_rois * n_time
print(f"  Total tests per model: {n_rois} × {n_time} = {total_tests}")
print(f"  FDR q = {FDR_Q}")
print("=" * 60)

explor_fdr_sig = np.zeros_like(roi_pvalues, dtype=bool)
for m_idx in range(n_models):
    p_matrix = roi_pvalues[:, :, m_idx]
    sig, threshold = fdr_bh(p_matrix, q=FDR_Q)
    explor_fdr_sig[:, :, m_idx] = sig
    n_sig = sig.sum()

    print(f"\n  {MODEL_LABELS[m_idx]}:")
    print(f"    FDR threshold: {threshold:.6f}")
    print(f"    Significant tests: {n_sig}/{total_tests}")

    if n_sig > 0:
        sig_indices = np.argwhere(sig)
        for si in sig_indices:
            r, t = si
            tw_s = time_windows[t, 0] * 1000
            tw_e = time_windows[t, 1] * 1000
            beta = roi_observed[r, t, m_idx]
            p = roi_pvalues[r, t, m_idx]
            print(f"      {roi_names[r]:>15s}  {tw_s:.0f}–{tw_e:.0f} ms  "
                  f"beta={beta:.4f}  p={p:.4f} ***")

# ----- 6. Save results -----
print(f"\nStep 6: Saving results to {OUTPUT_DIR}/")
np.save(os.path.join(OUTPUT_DIR, "roi_observed.npy"), roi_observed)
np.save(os.path.join(OUTPUT_DIR, "roi_pvalues.npy"), roi_pvalues)
np.save(os.path.join(OUTPUT_DIR, "explor_fdr_sig.npy"), explor_fdr_sig)
np.save(os.path.join(OUTPUT_DIR, "time_windows.npy"), time_windows)
np.save(os.path.join(OUTPUT_DIR, "perm_orders.npy"), perm_orders)
np.save(os.path.join(OUTPUT_DIR, "roi_perm_betas.npy"), roi_perm_betas)

with open(os.path.join(OUTPUT_DIR, "roi_info.pkl"), 'wb') as f:
    pickle.dump({
        'roi_names': roi_names,
        'roi_assignments': active_rois,
        'excluded_rois': excluded_rois,
        'model_names': MODEL_NAMES,
        'model_labels': MODEL_LABELS,
        'n_perm': N_PERM,
        'fdr_q': FDR_Q,
        'time_indices_global': time_indices.tolist(),
        'skip_time_idx': SKIP_TIME_IDX,
    }, f)

# ----- 7a. Heatmaps (separate SVGs for valence and arousal) -----
print("\nStep 7: Creating heatmaps")

tw_labels = [f"{time_windows[t,0]*1000:.0f}–{time_windows[t,1]*1000:.0f}"
             for t in range(n_time)]

for m_idx in range(n_models):
    model_label = MODEL_LABELS[m_idx]

    fig, ax = plt.subplots(figsize=(10, max(5, n_rois * 0.38)))

    data = roi_observed[:, :, m_idx]
    vmax = 0.1

    im = ax.imshow(data, aspect='auto', cmap='viridis',
                   vmin=-0.13, vmax=vmax, interpolation='nearest')

    # Mark significant cells
    for r in range(n_rois):
        for t in range(n_time):
            if explor_fdr_sig[r, t, m_idx]:
                ax.text(t, r, '*', ha='center', va='center',
                        fontsize=12, fontweight='bold', color='black')
            elif roi_pvalues[r, t, m_idx] < 0.05:
                ax.text(t, r, '·', ha='center', va='center',
                        fontsize=12, color='gray')

    ax.set_yticks(range(n_rois))
    ax.set_yticklabels(roi_names, fontsize=8)
    ax.set_xticks(range(n_time))
    ax.set_xticklabels(tw_labels, rotation=70, ha='right', fontsize=7)
    ax.set_xlabel('Time window (ms)', fontsize=11)
    ax.set_title(f'{model_label} — Partial RSA betas (exploratory)',
                 fontsize=13, fontweight='bold')
    plt.colorbar(im, ax=ax, shrink=0.6, label='Beta')

    plt.tight_layout()
    outfile = f"{OUTPUT_DIR}/heatmap_new_{model_label.lower()}.svg"
    fig.savefig(outfile, format='svg', dpi=300, bbox_inches='tight')
    print(f"  Saved: {outfile}")
    plt.close()


# ----- 7b. Time courses for MFG, ACC, PCC -----
print("\nStep 10: Saving individual ROI time courses")

COLOR_VALENCE = '#D62728'
COLOR_AROUSAL = '#1F77B4'
colors = [COLOR_VALENCE, COLOR_AROUSAL]

tw_mids = (time_windows[:, 0] + time_windows[:, 1]) / 2 * 1000

# Load permutation betas for CI bands
roi_perm_betas = np.load(os.path.join(OUTPUT_DIR, "roi_perm_betas.npy"))

for roi_name in ['MFG', 'ACC', 'PCC']:
    if roi_name not in roi_names:
        print(f"  {roi_name} not found in active ROIs, skipping.")
        continue

    r_idx = roi_names.index(roi_name)
    fig, ax = plt.subplots(figsize=(9, 3))

    for m_idx in range(n_models):
        betas = roi_observed[r_idx, :, m_idx]
        perm_se = np.std(roi_perm_betas[r_idx, :, m_idx, :], axis=-1)  # shape (n_time,)
        ci_low  = betas - 1.96 * perm_se
        ci_high = betas + 1.96 * perm_se

        # Shaded CI band
        ax.fill_between(tw_mids, ci_low, ci_high,
                        color=colors[m_idx], alpha=0.15, linewidth=0)

        # Main line
        ax.plot(tw_mids, betas, color=colors[m_idx], linewidth=1.5,
                label=MODEL_LABELS[m_idx], marker='o', markersize=5)

        # FDR-significant markers
        for t in range(n_time):
            if explor_fdr_sig[r_idx, t, m_idx]:
                ax.plot(tw_mids[t], betas[t], 'o', color=colors[m_idx],
                        markersize=8, markeredgecolor='black', markeredgewidth=1.5)

    ax.axhline(y=0, color='gray', linestyle='-', linewidth=0.5)
    ax.set_xlabel('Time (ms)', fontsize=11)
    ax.set_ylabel('Beta', fontsize=11)
    ax.legend(fontsize=9, loc='upper right', markerscale=2, handlelength=2)
    ax.tick_params(labelsize=9)

    plt.tight_layout()
    outfile = os.path.join(OUTPUT_DIR, f"timecourse_{roi_name}.svg")
    fig.savefig(outfile, format='svg', dpi=300, bbox_inches='tight')
    print(f"  Saved: {outfile}")
    plt.close()



# ----- 9. Glass brains for MFG, ACC, PCC -----
print("\n Saving glass brain plots")

try:
    from nilearn import plotting as nlplot
    import nibabel as nib

    roi_plot_config = {
        'MFG': {'color': COLOR_VALENCE, 'display_mode': 'z', 'cut_coords': [40]},
        'ACC': {'color': COLOR_VALENCE, 'display_mode': 'x', 'cut_coords': [0]},
        'PCC': {'color': COLOR_AROUSAL, 'display_mode': 'z', 'cut_coords': [-5]},
    }

    aal_data   = nib.load(AAL_ATLAS_FILE).get_fdata()
    aal_affine = nib.load(AAL_ATLAS_FILE).affine
    aal_shape  = aal_data.shape

    index_to_roi = {}
    for roi_name_key, label_substrings in ROIS.items():
        for substr in label_substrings:
            for i, label in enumerate(labels):
                if label == substr or substr in label:
                    idx = int(indices[i])
                    index_to_roi[idx] = roi_name_key

    for roi_name, cfg in roi_plot_config.items():
        if roi_name not in roi_names:
            print(f"  {roi_name} not in active ROIs, skipping.")
            continue

        # Build a binary mask NIfTI for this ROI
        mask = np.zeros(aal_shape, dtype=np.float32)
        for aal_idx, mapped_roi in index_to_roi.items():
            if mapped_roi == roi_name:
                mask[aal_data == aal_idx] = 1.0

        if mask.sum() == 0:
            print(f"  {roi_name}: empty mask, skipping.")
            continue

        roi_img = nib.Nifti1Image(mask, aal_affine)

        # Use a single-color colormap so the ROI renders in the desired color
        from matplotlib.colors import LinearSegmentedColormap
        cmap = LinearSegmentedColormap.from_list(
            'roi_cmap', ['white', cfg['color']], N=2)

        display = nlplot.plot_glass_brain(
            roi_img,
            display_mode=cfg['display_mode'],
            cmap=cmap,
            vmin=0,
            vmax=1,
            colorbar=False,
            annotate=False,
            draw_cross=False,
            alpha=0.85,
        )

        outfile = os.path.join(OUTPUT_DIR, f"glassbrain_{roi_name}.svg")
        display.savefig(outfile)
        display.close()
        print(f"  Saved: {outfile}")

except ImportError as e:
    print(f"  nilearn not available: {e}")

# ----- 9. Glass brains showing all ROIs colored by peak beta -----
print("\n  Creating glass brain figures (all ROIs, peak beta across time)...")

from nilearn.plotting import plot_glass_brain

for m_idx in range(n_models):
    model_label = MODEL_LABELS[m_idx]
    brain_vals = np.zeros(n_grid, dtype=np.float64)

    roi_labels = []
    for r_idx, roi_name in enumerate(roi_names):
        grid_idx = active_rois[roi_name]

        # Specific time window (e.g., t_local = 5 for 350-450 ms)
        SHOW_TIME = 1
        peak_beta = roi_observed[r_idx, SHOW_TIME, m_idx]
        
        brain_vals[grid_idx] = peak_beta
        roi_labels.append(f"{roi_name} (β={peak_beta:.3f})")

    if m_idx == 0:
        cmap = 'Reds'
    else:
        cmap = 'Blues'

    # Use absolute max across ROIs for consistent color scaling
    vmax = np.max(np.abs(brain_vals[brain_vals != 0]))
    vmax = max(vmax, 0.01)  # avoid zero range

    img = make_nifti(grid_mm, np.abs(brain_vals))  # abs so Reds/Blues work

    g = plot_glass_brain(
        img,
        display_mode="lzr",
        symmetric_cbar=False,
        colorbar=True,
        cmap=cmap,
        cbar_tick_format='%.3f',
        plot_abs=False,
        threshold=0.001,
        vmin=0,
        vmax=vmax,
    )

    outfile = f"{OUTPUT_DIR}/glass_brain_all_rois_{model_label.lower()}.svg"
    g.savefig(outfile, dpi=300)
    plt.close('all')
    print(f"  {model_label}: saved ({', '.join(roi_labels)})")
    
# ----- 10. Summary table -----
total_time = time.time() - t_start
print(f"\n{'=' * 60}")
print(f"DONE. Runtime: {total_time/60:.1f} min")
print(f"{'=' * 60}")

print(f"\nEXPLORATORY RESULTS TABLE")
print(f"{'ROI':>15s}  {'nPts':>5s}  {'Model':>10s}  {'Time (ms)':>12s}  "
      f"{'Beta':>8s}  {'p':>8s}  {'FDR':>5s}")
print(f"{'-'*15}  {'-'*5}  {'-'*10}  {'-'*12}  {'-'*8}  {'-'*8}  {'-'*5}")

for r_idx, roi_name in enumerate(roi_names):
    n_pts = len(active_rois[roi_name])
    for m_idx in range(n_models):
        for t in range(n_time):
            if roi_pvalues[r_idx, t, m_idx] < 0.05:
                tw_s = time_windows[t, 0] * 1000
                tw_e = time_windows[t, 1] * 1000
                beta = roi_observed[r_idx, t, m_idx]
                p = roi_pvalues[r_idx, t, m_idx]
                fdr = "YES" if explor_fdr_sig[r_idx, t, m_idx] else "no"
                print(f"{roi_name:>15s}  {n_pts:>5d}  {MODEL_LABELS[m_idx]:>10s}  "
                      f"{tw_s:5.0f}–{tw_e:5.0f}  {beta:>8.4f}  "
                      f"{p:>8.4f}  {fdr:>5s}")

# ----- 11. Save beta values to Excel -----
print("\nStep 7b: Saving beta values to Excel")

import openpyxl

wb = openpyxl.Workbook()

for m_idx, model_label in enumerate(MODEL_LABELS):
    ws = wb.active if m_idx == 0 else wb.create_sheet()
    ws.title = model_label

    # Header row: ROI names
    ws.cell(row=1, column=1, value="Time window (ms)")
    for r_idx, roi_name in enumerate(roi_names):
        ws.cell(row=1, column=r_idx + 2, value=roi_name)

    # Data rows
    for t in range(n_time):
        tw_label = f"{time_windows[t,0]*1000:.0f}–{time_windows[t,1]*1000:.0f}"
        ws.cell(row=t + 2, column=1, value=tw_label)
        for r_idx in range(n_rois):
            beta = roi_observed[r_idx, t, m_idx]
            ws.cell(row=t + 2, column=r_idx + 2,
                    value=None if np.isnan(beta) else round(float(beta), 4))

outfile_xlsx = os.path.join(OUTPUT_DIR, "roi_betas.xlsx")
wb.save(outfile_xlsx)
print(f"  Saved: {outfile_xlsx}")